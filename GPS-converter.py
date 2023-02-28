#installed packages: openpyxl,xlrd, pathlib2, tkintermapview, customtkinter
import tkinter
from tkinter import *
from tkinter import BOTH
from tkinter import END
import tkinter.font as font
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib2
import tkintermapview
from tkintermapview import TkinterMapView
from tkinter import IntVar, StringVar
from tkinter import messagebox
import customtkinter
import math
import re

customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("green")

root = customtkinter.CTk()
root.title('GPS Converter')
root.iconbitmap('logo-gps.ico')
root.geometry('1100x600+10+10')
root.resizable(0,0)

saved_coord = pathlib2.Path("Saved coordinates.xlsx")

if not saved_coord.exists():
    saved_coord=Workbook()
    sheet=saved_coord.active
    sheet['A1'] = 'DD Latitude'
    sheet['B1'] = 'DD Longitude'
    sheet['C1'] = 'DMS Latitude'
    sheet['D1'] = 'DMS Longitude'

    saved_coord.save("Saved coordinates.xlsx")

def convert_DDtoDMS():
    try:
        if latDDtoDMS.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DD latitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DD latitude value!')

    try:
        float(latDDtoDMS.get())
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DD latitude should be a number!')

    try:
        if float(latDDtoDMS.get())>90:
            return messagebox.showerror('Input value error', 'Error: DD latitude should be between -90 and 90!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: DD latitude should be between -90 and 90!')

    try:
        if float(latDDtoDMS.get())<-90:
            return messagebox.showerror('Input value error', 'Error: DD latitude should be between -90 and 90!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: DD latitude should be between -90 and 90!')

    try:
        if longDDtoDMS.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DD longitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DD longitude value!')

    try:
        float(longDDtoDMS.get())
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DD longitude should be a number!')

    try:
        if float(longDDtoDMS.get())>180:
            return messagebox.showerror('Input value error', 'Error: DD longitude should be between -180 and 180!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: DD longitude should be between -180 and 180!')

    try:
        if float(longDDtoDMS.get())<-180:
            return messagebox.showerror('Input value error', 'Error: DD longitude should be between -180 and 180!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: DD longitude should be between -180 and 180!')

    latitudeDMS = abs(float(latDDtoDMS.get()))
    whole_lat = math.trunc(latitudeDMS)
    remain1_lat = latitudeDMS - whole_lat
    min_lat = math.trunc(remain1_lat*60)
    remain2_lat = (remain1_lat*60) - min_lat
    sec_lat = round (remain2_lat*60,2)

    if latDDtoDMS.get().startswith('-'):
        DMS_lat = str(whole_lat) + '°' + str(min_lat) + "'" + str(sec_lat) + '"' + 'S'
    else:
        DMS_lat = str(whole_lat) + '°' + str(min_lat) + "'" + str(sec_lat) + '"' + 'N'

    longitudeDMS = abs(float(longDDtoDMS.get()))
    whole_long = math.trunc(longitudeDMS)
    remain1_long = longitudeDMS - whole_long
    min_long = math.trunc(remain1_long*60)
    remain2_long = (remain1_long*60) - min_long
    sec_long = round (remain2_long*60,2)
    if latDDtoDMS.get().startswith('-'):
        DMS_long = str(whole_long) + '°' + str(min_long) + "'" + str(sec_long) + '"' + 'W'
    else:
        DMS_long = str(whole_long) + '°' + str(min_long) + "'" + str(sec_long) + '"' + 'E'

    global latitude_DMS_output
    latitude_DMS_output = tkinter.Label(convert_frame,text=DMS_lat, font=("Roboto Medium", 12), width= 20)
    latitude_DMS_output.grid(row=1, column=4, padx=5, pady=5)
    global longitude_DMS_output
    longitude_DMS_output = tkinter.Label(convert_frame, text = DMS_long,font=("Roboto Medium", 12), width= 20)
    longitude_DMS_output.grid(row=2, column=4, padx=5, pady=5)

    map_widget_sat = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
    map_widget_sat.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    x = float(latDDtoDMS.get())
    y = float(longDDtoDMS.get())
    map_widget_sat.set_position(x, y, marker=True)
    map_widget_sat.set_zoom(15)
    map_widget_sat.grid(row=0, column=0, padx=5, pady=5)
    map_widget_norm = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
    map_widget_norm.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    map_widget_norm.set_position(x, y, marker=True)
    map_widget_norm.set_zoom(15)
    map_widget_norm.grid(row=0, column=1, padx=5, pady=5)

    save_button_DD['state'] = tkinter.NORMAL
    save_button_DD_disab1.grid_forget()

def reset_DDtoDMS():
    latDDtoDMS.delete(0, END)
    latitude_DMS_output.config(text="")
    longDDtoDMS.delete(0, END)
    longitude_DMS_output.config(text="")

def save_DDtoDMS():
    try:
        if latDDtoDMS.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DD latitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DD latitude value!')

    try:
        if longDDtoDMS.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DD longitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DD longitude value!')

    try:
        if (latitude_DMS_output['text'] == ''):
            return messagebox.showerror('Input value error', 'Error: Please, convert the data!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, convert the data!')

    latDD1=latDDtoDMS.get()
    latDMS1=latitude_DMS_output.cget('text')
    longDD1=longDDtoDMS.get()
    longDMS1=longitude_DMS_output.cget('text')
    print(latDD1)
    print(longDD1)
    print(latDMS1)
    print(longDMS1)

    saved_coord=openpyxl.load_workbook("Saved coordinates.xlsx")
    sheet=saved_coord.active
    sheet.cell(column=1,row=sheet.max_row+1,value=latDD1)
    sheet.cell(column=2,row=sheet.max_row,value=longDD1)
    sheet.cell(column=3,row=sheet.max_row,value=latDMS1)
    sheet.cell(column=4,row=sheet.max_row,value=longDMS1)
    saved_coord.save("Saved coordinates.xlsx")

    popup1 = customtkinter.CTkToplevel(root)
    popup1.title('Confirmation message')
    popup1.geometry('300x50')
    pop1label = customtkinter.CTkLabel(popup1, text="Saved successfully", text_font=('Roboto Medium',11))
    pop1label.pack(fill='x', padx=5, pady=5)
    popup1.after(1500, lambda: popup1.destroy())

    save_button_DD['state'] = tkinter.DISABLED
    global save_button_DD_disab1
    save_button_DD_disab1 = tkinter.Button(convert_frame, text='Save to Excel', font=('Roboto Medium',12), command= save_DDtoDMS_message1)
    save_button_DD_disab1.grid(row=2, column=6)

def save_DDtoDMS_message1():
    if (save_button_DD['state'] == tkinter.DISABLED):
        popup1 = customtkinter.CTkToplevel(root)
        popup1.geometry('350x50')
        popup1.title('Error message')
        pop1label = customtkinter.CTkLabel(popup1, text="Convert new data to activate Save button", text_font=('Roboto Medium',11))
        pop1label.pack(fill='x', padx=5, pady=5)
        popup1.after(2000, lambda: popup1.destroy())

def convert_DMStoDD():
    latitudeDD = re.split('[°,\',\"]',latDMStoDD.get())

    try:
        if latDMStoDD.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DMS latitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DMS latitude value!')

    try:
        float(latitudeDD[0])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS latitude should contain only numbers!')

    try:
        float(latitudeDD[1])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS latitude should contain only numbers!')

    try:
        float(latitudeDD[2])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS latitude should contain only numbers!')

    if '' in latitudeDD:
        pass
    else:
        messagebox.showerror('Input value error', 'Error: DMS latitude should be entered in the format XX°XX\'XX\"!')

    try:
        if float(latitudeDD[0])>90:
            return messagebox.showerror('Input value error', 'Error: Latitude degrees should be between 0 and 90!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude degrees should be between 0 and 90!')

    try:
        if float(latitudeDD[0])<0:
            return messagebox.showerror('Input value error', 'Error: Latitude degrees should be between 0 and 90!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude degrees should be between 0 and 90!')

    try:
        if float(latitudeDD[1])>60:
            return messagebox.showerror('Input value error', 'Error: Latitude minutes should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude minutes should be between 0 and 60!')

    try:
        if float(latitudeDD[1])<0:
            return messagebox.showerror('Input value error', 'Error: Latitude minutes should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude minutes should be between 0 and 60!')

    try:
        if float(latitudeDD[2])>60:
            return messagebox.showerror('Input value error', 'Error: Latitude seconds should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude seconds should be between 0 and 60!')

    try:
        if float(latitudeDD[2])<0:
            return messagebox.showerror('Input value error', 'Error: Latitude seconds should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Latitude seconds should be between 0 and 60!')

    longitudeDD = re.split('[°,\',\"]',longDMStoDD.get())

    try:
        if longDMStoDD.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DMS longitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DMS longitude value!')

    try:
        float(longitudeDD[0])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS longitude should contain only numbers!')

    try:
        float(longitudeDD[1])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS longitude should contain only numbers!')

    try:
        float(longitudeDD[2])
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: DMS longitude should contain only numbers!')

    if '' in longitudeDD:
        pass
    else:
        messagebox.showerror('Input value error', 'Error: DMS longitude should be entered in the format XX°XX\'XX\"!')

    try:
        if float(longitudeDD[0])>180:
            return messagebox.showerror('Input value error', 'Error: Longitude degrees should be between 0 and 180!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude degrees should be between 0 and 180!')

    try:
        if float(longitudeDD[0])<0:
            return messagebox.showerror('Input value error', 'Error: Longitude degrees should be between 0 and 180!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude degrees should be between 0 and 180!')

    try:
        if float(longitudeDD[1])>60:
            return messagebox.showerror('Input value error', 'Error: Longitude minutes should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude minutes should be between 0 and 60!')

    try:
        if float(longitudeDD[1])<0:
            return messagebox.showerror('Input value error', 'Error: Longitude minutes should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude minutes should be between 0 and 60!')

    try:
        if float(longitudeDD[2])>60:
            return messagebox.showerror('Input value error', 'Error: Longitude seconds should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude seconds should be between 0 and 60!')

    try:
        if float(longitudeDD[2])<0:
            return messagebox.showerror('Input value error', 'Error: Longitude seconds should be between 0 and 60!')
    except ValueError:
        messagebox.showerror('Input value error', 'Error: Longitude seconds should be between 0 and 60!')

    latitudeDD.remove("")

    if 2 < len(latitudeDD):
        d_lat = float(latitudeDD[0])
        m_lat = float(latitudeDD[1])/60
        s_lat = float(latitudeDD[2])/3600
        dms_lat = d_lat+m_lat+s_lat
    elif 1 < len(latitudeDD):
        d_lat = float(latitudeDD[0])
        m_lat = float(latitudeDD[1])/60
        dms_lat = d_lat+m_lat
    elif 0 < len(latitudeDD):
        d_lat = float(latitudeDD[0])
        dms_lat = d_lat
    else:
        messagebox.showerror('Input value error', 'Error: DMS latitude should be entered in the format XX°XX\'XX\"!')

    if directionNS.get() == '1':
        dms_direction_lat = "{:.5f}".format(dms_lat)
    elif directionNS.get() == '-1':
        dms_direction_lat = "{:.5f}".format(dms_lat*-1)
    else:
        messagebox.showerror('Input value error', 'Error: Please, select direction (N or S)!')

    global latitude_DD_output
    latitude_DD_output = tkinter.Label(convert_frame, text = dms_direction_lat, font=("Roboto Medium", 12),width= 20)
    latitude_DD_output.grid(row=4, column=4, padx=5, pady=5)

    longitudeDD.remove("")

    if 2 < len(longitudeDD):
        d_long = float(longitudeDD[0])
        m_long = float(longitudeDD[1])/60
        s_long = float(longitudeDD[2])/3600
        dms_long = d_long+m_long+s_long
    elif 1 < len(longitudeDD):
        d_long = float(longitudeDD[0])
        m_long = float(longitudeDD[1])/60
        dms_long = d_long+m_long
    elif 0 < len(longitudeDD):
        d_long = float(longitudeDD[0])
        dms_long = d_long
    else:
        messagebox.showerror('Input value error', 'Error: DMS longitude should be entered in the format XX°XX\'XX\"!')

    if directionEW.get() == '1':
        dms_direction_long = "{:.5f}".format(dms_long)
    elif directionEW.get() == '-1':
        dms_direction_long = "{:.5f}".format(dms_long*-1)
    else:
        messagebox.showerror('Input value error', 'Error: Please, select direction (E or W)!')

    global longitude_DD_output
    longitude_DD_output = tkinter.Label(convert_frame, text = dms_direction_long, font=("Roboto Medium", 12),width= 20)
    # longitude_DD_output.grid(row=5, column=4, padx=5, pady=5)
    longitude_DD_output.grid(row=5, column=4)

    map_widget_sat = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
    map_widget_sat.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    x = float(latitude_DD_output.cget('text'))
    y = float(longitude_DD_output.cget('text'))
    map_widget_sat.set_position(x, y, marker=True)
    map_widget_sat.set_zoom(15)
    map_widget_sat.grid(row=0, column=0, padx=5, pady=5)
    map_widget_norm = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
    map_widget_norm.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
    map_widget_norm.set_position(x, y, marker=True)
    map_widget_norm.set_zoom(15)
    map_widget_norm.grid(row=0, column=1, padx=5, pady=5)

    save_button_DMS['state'] = tkinter.NORMAL
    save_button_DD_disab2.grid_forget()

def reset_DMStoDD():
    longDMStoDD.delete(0, END)
    longitude_DD_output.config(text="")
    latDMStoDD.delete(0, END)
    latitude_DD_output.config(text="")
    directionNS.set(None)
    directionEW.set(None)

def save_DMStoDD():
    try:
        if latDMStoDD.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DMS latitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DMS latitude value!')

    try:
        if longDMStoDD.get() == '':
            return messagebox.showerror('Input value error', 'Error: Please, enter the DMS longitude value!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, enter the DMS longitude value!')

    try:
        if (latitude_DD_output['text'] == ''):
            return messagebox.showerror('Input value error', 'Error: Please, convert the data!')
    except ValueError:
        return messagebox.showerror('Input value error', 'Error: Please, convert the data!')

    latDMS2=latDMStoDD.get()
    latDD2=latitude_DD_output.cget('text')
    longDMS2=longDMStoDD.get()
    longDD2=longitude_DD_output.cget('text')
    print(latDD2)
    print(latDMS2)
    print(longDD2)
    print(longDMS2)

    saved_coord=openpyxl.load_workbook("Saved coordinates.xlsx")
    sheet=saved_coord.active
    sheet.cell(column=1,row=sheet.max_row+1,value=latDD2)
    sheet.cell(column=2,row=sheet.max_row,value=longDD2)
    sheet.cell(column=3,row=sheet.max_row,value=latDMS2)
    sheet.cell(column=4,row=sheet.max_row,value=longDMS2)
    saved_coord.save("Saved coordinates.xlsx")

    popup2 = customtkinter.CTkToplevel(root)
    popup2.geometry('300x50')
    popup2.title('Confirmation message')
    pop2label = customtkinter.CTkLabel(popup2, text="Saved successfully", text_font=('Roboto Medium',11))
    pop2label.pack(fill='x', padx=5, pady=5)
    popup2.after(1500, lambda: popup2.destroy())

    save_button_DMS['state'] = tkinter.DISABLED
    global save_button_DD_disab2
    save_button_DD_disab2 = tkinter.Button(convert_frame, text='Save to Excel', font=('Roboto Medium',12), command= save_DMStoDD_message2)
    save_button_DD_disab2.grid(row=5, column=6)

def save_DMStoDD_message2():
    if (save_button_DMS['state'] == tkinter.DISABLED):
        popup2 = customtkinter.CTkToplevel(root)
        popup2.geometry('350x50')
        popup2.title('Error message')
        pop2label = customtkinter.CTkLabel(popup2, text="Convert new data to activate Save button", text_font=('Roboto Medium',11))
        pop2label.pack(fill='x', padx=5, pady=5)
        popup2.after(2000, lambda: popup2.destroy())

convert_frame = customtkinter.CTkFrame(root)
label_frame = customtkinter.CTkFrame(root, width=1100, height=50)
map_frame = customtkinter.CTkFrame(root)

convert_frame.pack(fill=BOTH, expand = False)
label_frame.pack()
map_frame.pack(fill=BOTH, expand = True)

customtkinter.CTkLabel(convert_frame, text='DD (decimal degrees)', text_font=("Roboto Medium", 12)).grid(row=0, column=1, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='DMS (degrees, minutes, seconds)', text_font=("Roboto Medium", 12)).grid(row=0, column=4, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='Latitude', text_font=("Roboto Medium", 12)).grid(row=1, column=0, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='Longitude', text_font=("Roboto Medium", 12)).grid(row=2, column=0, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='DMS (degrees, minutes, seconds)', text_font=("Roboto Medium", 12)).grid(row=3, column=1, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='DD (decimal degrees)', text_font=("Roboto Medium", 12)).grid(row=3, column=4, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='Latitude', text_font=("Roboto Medium", 12)).grid(row=4, column=0, padx=5, pady=5)
customtkinter.CTkLabel(convert_frame, text='Longitude', text_font=("Roboto Medium", 12)).grid(row=5, column=0, padx=5, pady=5)

latitude_DMS_output = tkinter.Label(convert_frame,text='', width= 20)
latitude_DMS_output.grid(row=1, column=4, padx=5, pady=5)
longitude_DMS_output = tkinter.Label(convert_frame,text = '', width= 20)
longitude_DMS_output.grid(row=2, column=4, padx=5, pady=5)
longitude_DD_output = tkinter.Label(convert_frame, text = '', width= 20)
longitude_DD_output.grid(row=5, column=4, padx=5, pady=5)
latitude_DD_output = tkinter.Label(convert_frame, text = '', width= 20)
latitude_DD_output.grid(row=4, column=4, padx=5, pady=5)

latDDtoDMS = customtkinter.CTkEntry(convert_frame, text_font=("Roboto Medium", 12),width=150)
latDDtoDMS.grid(row=1, column=1, padx=5, pady=5)
longDDtoDMS = customtkinter.CTkEntry(convert_frame, text_font=("Roboto Medium", 12),width=150)
longDDtoDMS.grid(row=2, column=1, padx=5, pady=5)
latDMStoDD = customtkinter.CTkEntry(convert_frame, text_font=("Roboto Medium", 12),width=150)
latDMStoDD.grid(row=4, column=1, padx=5, pady=5)
longDMStoDD = customtkinter.CTkEntry(convert_frame, text_font=("Roboto Medium", 12),width=150)
longDMStoDD.grid(row=5, column=1, padx=5, pady=5)

directionNS = StringVar()
directionEW = StringVar()

directionNS.set(None)
directionEW.set(None)

N = customtkinter.CTkRadioButton(convert_frame, text = 'N', text_font=("Roboto Medium", 12), variable=directionNS, value = '1')
S = customtkinter.CTkRadioButton(convert_frame, text = 'S', text_font=("Roboto Medium", 12), variable=directionNS, value = '-1')
E = customtkinter.CTkRadioButton(convert_frame, text = 'E', text_font=("Roboto Medium", 12),variable=directionEW, value = '1')
W = customtkinter.CTkRadioButton(convert_frame, text = 'W', text_font=("Roboto Medium", 12),variable=directionEW, value = '-1')

N.grid(row=4,column=2, padx=5, pady=5)
S.grid(row=4,column=3, padx=5, pady=5)
E.grid(row=5,column=2, padx=5, pady=5)
W.grid(row=5,column=3, padx=5, pady=5)

convert_button_DD = customtkinter.CTkButton(convert_frame, text='Convert DD to DMS', text_font=("Roboto Medium", 12), command= convert_DDtoDMS)
reset_button_DD = customtkinter.CTkButton(convert_frame, text='Reset', text_font=("Roboto Medium", 12),command=reset_DDtoDMS)
save_button_DD = tkinter.Button(convert_frame, text='Save to Excel', font=("Roboto Medium", 12),command= save_DDtoDMS)

convert_button_DMS = customtkinter.CTkButton(convert_frame, text='Convert DMS to DD', text_font=("Roboto Medium", 12),command= convert_DMStoDD)
reset_button_DMS = customtkinter.CTkButton(convert_frame, text='Reset', text_font=("Roboto Medium", 12),command=reset_DMStoDD)
save_button_DMS = tkinter.Button(convert_frame, text='Save to Excel', font=("Roboto Medium", 12),command= save_DMStoDD)

convert_button_DD.grid(row=1, column=5, padx=5, pady=5)
reset_button_DD.grid(row=1, column=6, padx=5, pady=5)
save_button_DD.grid(row=2, column=6, padx=5, pady=5)
convert_button_DMS.grid(row=4, column=5, padx=5, pady=5)
reset_button_DMS.grid(row=4, column=6, padx=5, pady=5)
save_button_DMS.grid(row=5, column=6, padx=5, pady=5)

map_label =customtkinter.CTkLabel(label_frame, text='Searched localisation on map', text_font=("Roboto Medium", 12)).place(anchor = CENTER, relx = .5, rely = .8)

map_widget_sat = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
map_widget_sat.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
map_widget_sat.set_position(50.0612, 19.9377, marker=True)
map_widget_sat.set_zoom(15)
map_widget_sat.grid(row=0, column=0, padx=5, pady=5)
map_widget_norm = tkintermapview.TkinterMapView(map_frame, width=815, height=400)
map_widget_norm.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
map_widget_norm.set_position(50.0612, 19.9377, marker=True)
map_widget_norm.set_zoom(15)
map_widget_norm.grid(row=0, column=1, padx=5, pady=5)

customtkinter.CTkLabel(map_frame, text='Satellite Google map', text_font=("Roboto Medium", 12)).grid(row=1, column=0, padx=2, pady=2)
customtkinter.CTkLabel(map_frame, text='Normal Google map', text_font=("Roboto Medium", 12)).grid(row=1, column=1, padx=2, pady=2)

root.mainloop()
